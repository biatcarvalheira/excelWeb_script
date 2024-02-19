import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from pandas.tseries.offsets import BMonthBegin
from openpyxl.styles import NamedStyle
import openpyxl
from excel_parser.excel_parser import project_root, process_xlsx_trade, extract_underlying_symbol, process_xlsx_orders
from web_scraper.web_scraper import run_all_web


def main():
    c = menu()
    df_trade = ''
    df_orders = ''
    if c == '1' or c == '3':
        c_list = process_xlsx_trade(data_input_directory_trade)
        list_type = 'trade'
        df_trade = format_data(column_headers, c_list, data_input_directory_trade, c, list_type)
    if c == '2' or c == '3':
        orders_c_list = process_xlsx_orders(data_input_directory_orders)
        list_type = 'orders'
        df_orders = format_data(column_headers, orders_c_list, data_input_directory_orders, c, list_type)
    if c == '1':
        save_data(df_trade, data_output_directory_trade)
        insert_line_after(data_output_directory_trade, 'Sheet1', 1, first_line_data)
        # format percentage cells
        format_columns(data_output_directory_trade, 'Sheet1', ['L', 'N', 'Q', 'X', 'Y', 'AB'], 'percentage', '0.00%')
        # format currency cells
        format_columns(data_output_directory_trade, 'Sheet1',
                       ['T', 'V', 'AI', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS'], 'currency_format',
                       '"$"#,##0.00')
    if c == '2':
        save_data(df_orders, data_output_directory_orders)
        insert_line_after(data_output_directory_orders, 'Sheet1', 1, first_line_data)
        # format percentage cells
        format_columns(data_output_directory_orders, 'Sheet1', ['L', 'N', 'X', 'Y', 'AB'], 'percentage',
                       '0.00%')
        # format currency cells
        format_columns(data_output_directory_orders, 'Sheet1',
                       ['T', 'V', 'AI', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS'], 'currency_format',
                       '"$"#,##0.00')


    if c == '3':
        save_multiple_dataframes(df_trade, df_orders, data_output_directory_trade_and_orders)
        insert_line_after(data_output_directory_trade_and_orders, 'Sheet1', 1, first_line_data)
        # format percentage cells
        format_columns(data_output_directory_trade_and_orders, 'Sheet1', ['L', 'N', 'Q', 'X', 'Y', 'AB'], 'percentage', '0.00%')
        # format currency cells
        format_columns(data_output_directory_trade_and_orders, 'Sheet1',
                      ['T', 'V', 'AI', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS'], 'currency_format',
                       '"$"#,##0.00')

        insert_line_after(data_output_directory_trade_and_orders, 'Sheet2', 1, first_line_data)
        # format percentage cells
        format_columns(data_output_directory_trade_and_orders, 'Sheet2', ['L', 'N', 'X', 'Y', 'AB'], 'percentage',
                       '0.00%')
        # format currency cells
        format_columns(data_output_directory_trade_and_orders, 'Sheet2',
                       ['T', 'V', 'AI', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS'], 'currency_format',
                       '"$"#,##0.00')


def menu():
    print('Menu:')
    print('1. Run Trade Mixer')
    print('2. Run Orders Mixer')
    print('3. Run Trade and Orders Mixer')
    print('0. Exit')
    choice = input('(Enter your choice (0-3):')
    return choice


def format_data(column_h, content_list, data_input, choice_number, type_of_list):
    und_symbol = extract_underlying_symbol(data_input)
    uptt_list, mkt_b_list = run_all_web(und_symbol)
    number_of_headers = len(column_h)

    # Create an empty DataFrame with 40 columns
    columns = [column_h[i] for i in range(1, number_of_headers)]
    df = pd.DataFrame(columns=columns)
    df_list = ['']
    if type_of_list == 'trade':
        df_list = ['trade date', 'option Expiration date', 'days till exp (trade date)', 'days till exp (current)',
                   'Strike', 'underlying symbol', 'underlying price at time of trade', 'otm at time of trade',
                   'underlying price, current', 'otm, current', '$ amount of stock itm can be called (-) or put (+)',
                   'weight', 'weighted otm', 'mkt beta', 'Type', 'mkt beta* mkt px*contracts', 'Qty',
                   'mkt price *number of contracts', 'Trade Price/premium', 'trade price as percent of notional',
                   'annual yield at strike at time of trade', 'yield at current mkt price at time of trade', 'premium',
                   f'contracted in {previous_5_months[3]}',
                   f'contracted in {previous_5_months[2]}', f'contracted in {previous_5_months[1]}',
                   f'contracted in {previous_5_months[0]}', 'cash if exercised', '=AK1-A1', '=AL1-A1', '=AM1-A1', '=AN1-A1',
                   '=AO1-A1', '=AP1-A1', '=AQ1-A1', '=AR1-A1', '=AS1-A1', '=AT1-A1', '=AU1-A1'
                   ]
    elif type_of_list == 'orders':
        df_list = ['trade date', 'option Expiration date', 'days till exp (trade date)', 'days till exp (current)', 'order expiration date "time in force"',
                   'Strike', 'underlying symbol', 'underlying price at time of trade', 'otm at time of trade',
                   'underlying price, current', 'otm, current', '$ amount of stock itm can be called (-) or put (+)', 'Type', 'Qty',
                   'Trade Price/premium', 'trade price as percent of notional',
                   'annual yield at strike at time of trade', 'yield at current mkt price at time of trade', 'premium',
                   f'contracted in {previous_5_months[3]}',
                   f'contracted in {previous_5_months[2]}', f'contracted in {previous_5_months[1]}',
                   f'contracted in {previous_5_months[0]}', 'cash if exercised', '=AK1-A1', '=AL1-A1', '=AM1-A1',
                   '=AN1-A1',
                   '=AO1-A1', '=AP1-A1', '=AQ1-A1', '=AR1-A1', '=AS1-A1', '=AT1-A1', '=AU1-A1'
                   ]
    for index, c in enumerate(content_list):
        try:
           # print('length of content list item', len(c))
            if df_list[index] == 'underlying price at time of trade' and len(c) <= 3 and (choice_number == '1' or choice_number == '3'):
                c = uptt_list
            if df_list[index] == 'mkt beta' and len(c) <= 3 and (choice_number == '1' or choice_number == '3'):
                c = mkt_b_list
            df[df_list[index]] = c
        except IndexError:
            # Print the index of the list that caused the error
            print(f"List at index {index} is out of range.")

    # Fill specific columns with initial values
    df.insert(0, 'check date >>', '')  # or use an empty string: ''

    return df
    # Specify the Excel file path


def save_data(data_frame, saving_directory):
    # Save the DataFrame to an Excel file
    data_frame.to_excel(saving_directory, index=False)
    print(f'Data saved to {saving_directory}')
    print('************ Processes completed successfully! ************')

def save_multiple_dataframes(df1, df2, saving_directory, sheet1_name='Sheet1', sheet2_name='Sheet2'):
    # Use ExcelWriter to save to a specific sheet
    with pd.ExcelWriter(saving_directory, mode='w', engine='openpyxl') as writer:
        # Save the first DataFrame to the first sheet
        df1.to_excel(writer, sheet_name=sheet1_name, index=False)

        # Save the second DataFrame to the second sheet
        df2.to_excel(writer, sheet_name=sheet2_name, index=False)

    print(f'Data saved to {saving_directory} in sheets "{sheet1_name}" and "{sheet2_name}"')
    print('************ Processes completed successfully! ************')

def insert_line_after(file_path, sheet_name, row_number, data):
    # Load the existing workbook
    wb = load_workbook(file_path)

    # Select the desired sheet
    sheet = wb[sheet_name]

    # Shift existing rows down to make space for the new line
    sheet.insert_rows(row_number)

    # Write the data to the new line
    for col_num, value in enumerate(data, start=1):
        sheet.cell(row=row_number, column=col_num, value=value)

    # Save the changes
    wb.save(file_path)


def format_columns(file_path, sheet_name, column_letters, formatting_name, formatting_number):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)

    # Select the desired sheet
    sheet = workbook[sheet_name]

    # Define the starting row
    starting_row = 3  # Change this to the row where formatting should begin

    # Convert column letters to column indices
    column_indices = [openpyxl.utils.column_index_from_string(col) for col in column_letters]

    # Create a custom style with the specified formatting
    formatting_style = NamedStyle(name=formatting_name, number_format=formatting_number)

    # Apply the custom style to the specified range in each column
    for col_index in column_indices:
        for row in sheet.iter_rows(min_col=col_index, max_col=col_index, min_row=starting_row):
            for cell in row:
                cell.style = formatting_style

    # Save the changes to the Excel file
    workbook.save(file_path)


# --- Next 9 Fridays Dates --- #
def find_next_11_fridays():
    # Get today's date
    today = datetime.now().date()

    # Find the next Friday from today
    days_until_next_friday = (4 - today.weekday() + 7) % 7
    next_friday = today + timedelta(days=days_until_next_friday)

    # Calculate the dates for the next 10 Fridays
    next_fridays_primary_list = [next_friday + timedelta(weeks=i) for i in range(11)]
    next_fridays = []
    # Print the result in "mm/dd/yy" format
    for date in next_fridays_primary_list:
        try:
            formatted_date = date.strftime("%-m/%d/%y")
            next_fridays.append(formatted_date)
        except IndexError:
            # Print the index of the list that caused the error
            print(f"List at index {date} is out of range.")

    return next_fridays


# --- Last Five Months --- #
def last_five_months_writing():
    today = datetime.now()
    last_five_months_dates = [today - timedelta(days=30 * i) for i in range(4)]

    # Formatting the month names and printing in reverse order
    formatted_months = [date.strftime('%B') for date in last_five_months_dates][::-1]

    return formatted_months



# --- LISTS AND OTHER DATA --- #

# --- First Business Day (Cell2) --- #
# Get the current date
current_date = datetime.now()

# first business day
first_business_day_current_month = pd.date_range(start=current_date, periods=1, freq=BMonthBegin()).normalize()[0]

# Format the result to mm/dd/yy without the time
formatted_result = first_business_day_current_month.strftime('%m/%d/%y')

date_string = formatted_result
date_format = "%m/%d/%y"

# Convert string to date
date_object = datetime.strptime(date_string, date_format)

# Format date as a string
formatted_date = date_object.strftime('%m/%d/%y')

# --- Date of Extraction --- #
timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
header_time_stamp = datetime.now().strftime('%m/%d/%y')

# Todays date
today = datetime.now().date()

# --- Previous Months function usage --- #
previous_5_months = last_five_months_writing()


# -- formatting styles --#
percentage_style = NamedStyle(name='percentage', number_format='0.00%')
date_style = NamedStyle(name='date', number_format='d-mmm-yyyy')

# ------------------#
trade_name = 'TDA_YAHOO_CNBC_DATA_TRADE_DATA_'
orders_name = 'TDA_YAHOO_CNBC_DATA_ORDERS_DATA_'
trade_and_orders = 'TDA_YAHOO_CNBC_DATA_TRADE_AND_ORDERS_DATA_'

excel_filename_trade = f'{trade_name}{timestamp}.xlsx'
excel_filename_orders = f'{orders_name}{timestamp}.xlsx'
excel_filename_trade_and_orders = f'{trade_and_orders}{timestamp}.xlsx'

#### ---- Executable ---- #####
data_output_directory_trade = os.path.join(project_root, "data", "output", excel_filename_trade)
data_output_directory_orders = os.path.join(project_root, "data", "output", excel_filename_orders)
data_output_directory_trade_and_orders = os.path.join(project_root, "data", "output", excel_filename_trade_and_orders)

####---- IDE ---- #####
#data_output_directory_trade = os.path.join(project_root, "excelWeb_script", "data", "output", excel_filename_trade)

#data_output_directory_orders = os.path.join(project_root, "excelWeb_script", "data", "output", excel_filename_orders)

#data_output_directory_trade_and_orders = os.path.join(project_root, "excelWeb_script", "data", "output", excel_filename_trade_and_orders)


# ---- Executable ---- #
# --trade -- #
data_input_directory_trade = os.path.join(project_root, "data", "input", "trade")

# --orders -- #
data_input_directory_orders = os.path.join(project_root, "data", "input", "orders")


# ---- IDE ---- #
### --trade -- ####
#data_input_directory_trade = os.path.join(project_root, "excelWeb_script", "data", "input", "trade")

### --orders -- ####
#data_input_directory_orders = os.path.join(project_root, "excelWeb_script", "data", "input", "orders")



column_headers = [
    'check date >>', header_time_stamp, 'trade date', 'option Expiration date', 'days till exp (trade date)',
    'days till exp (current)', 'order expiration date "time in force"', 'days till expiration (if an order)',
    'Strike', 'underlying symbol',
    'underlying price at time of trade', 'otm at time of trade', 'underlying price, current', 'otm, current',
    '$ amount of stock itm can be called (-) or put (+)', 'weight', 'weighted otm', 'mkt beta', 'Type',
    'mkt beta* mkt px*contracts', 'Qty',
    'mkt price *number of contracts', 'Trade Price/premium', 'trade price as percent of notional',
    'annual yield at strike at time of trade', 'yield on cost at time of trade', 'multiple on cost',
    'yield at current mkt price at time of trade', 'premium', f'contracted in {previous_5_months[0]}',
    f'contracted in {previous_5_months[1]}', f'contracted in {previous_5_months[2]}',
    f'contracted in {previous_5_months[3]}', 'cash if exercised', 'days >>',
    '=AK1-A1', '=AL1-A1', '=AM1-A1', '=AN1-A1', '=AO1-A1', '=AP1-A1', '=AQ1-A1', '=AR1-A1', '=AS1-A1', '=AT1-A1',
    '=AU1-A1'
]

# first line content
first_line_data = [header_time_stamp, formatted_date, 'Open Positions', "Total"] + [None] * (len(column_headers) - 3)
fridays_list = find_next_11_fridays()
start_position = 35
for f in fridays_list:
    first_line_data[start_position] = f
    start_position += 1
try:
    main()

except Exception as e:
    print(f'Main. Error loading the program. {e}\nPlease try again.')
