import os
import datetime as datetime
import openpyxl
import sys
import re
from datetime import datetime, timedelta

# from src.web_scraper.web_scraper import underlying_price_at_time_of_trade
# from src.web_scraper.web_scraper import mkt_beta_list

# Get the absolute path to the script
script_path = os.path.abspath(sys.argv[0])

# Get the directory where the script is located (folder containing scripts)
script_directory = os.path.dirname(script_path)

# ---- to use when working in IDE ---- #
#project_root = os.path.abspath(os.path.join(script_directory, "..", ".."))


# ---- to use when exporting as an executable --- #
project_root = os.path.abspath(os.path.join(script_directory))

def get_xlsx(directory_path):
    try:
        # Get a list of all files in the directory
        files = os.listdir(directory_path)
        # Check if any of the files has a .xlsx extension
        for file in files:
            if file.lower().endswith('.xlsx') and not file.startswith('~$'):
                file_path = os.path.join(directory_path, file)
                try:
                    wb = openpyxl.load_workbook(file_path)
                    sheet_names = wb.sheetnames
                    if sheet_names:
                        first_sheet_name = sheet_names[0]
                        sheet = wb[first_sheet_name]

                        # Get the letters of all active columns
                        active_columns_letters = [openpyxl.utils.get_column_letter(col) for col in
                                                  range(1, sheet.max_column + 1)]

                        # Get the last active row
                        last_active_row = sheet.max_row

                        return first_sheet_name, file_path, active_columns_letters, last_active_row
                except Exception as e:
                    print(f"Error while processing file '{file}': {e}")
        # If no XLSX file with a sheet is found
        return None
    except Exception as e:
        print(f"Trade Excel processing error: error while listing directory '{directory_path}': {e}")


def remove_first_row_from_sheet(sheet):
    # Remove the first row
    removed_row = sheet[1]
    sheet.delete_rows(1)
    return removed_row


def get_xlsx_orders(directory_path):
    try:
        # Get a list of all files in the directory
        files = os.listdir(directory_path)
        # Check if any of the files has a .xlsx extension
        for file in files:
            if file.lower().endswith('.xlsx') and not file.startswith('~$'):
                file_path = os.path.join(directory_path, file)
                try:
                    wb = openpyxl.load_workbook(file_path)
                    sheet_names = wb.sheetnames
                    if sheet_names:
                        first_sheet_name = sheet_names[0]
                        sheet = wb[first_sheet_name]
                        removed_row = remove_first_row_from_sheet(sheet)

                        active_columns_letters = [openpyxl.utils.get_column_letter(col) for col in
                                                  range(1, sheet.max_column + 1)]

                        # Get the last active row
                        last_active_row = sheet.max_row
                        return first_sheet_name, file_path, active_columns_letters, last_active_row

                except Exception as e:
                    print(f"Error while processing file '{file}': {e}")
    except Exception as e:
        print(f"Trade Excel processing error: error while listing directory '{directory_path}': {e}")


def get_data_from_range(file_path, sheet_name, column_letters, start_row, end_row):
    try:
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(file_path)

        # Get the desired worksheet
        worksheet = workbook[sheet_name]

        # Initialize a list to store the separate column lists
        num_columns = len(column_letters)
        list_of_lists = [[] for _ in range(num_columns)]

        # Get header values excluding 'none'
        header_row = worksheet[1]
        header_values = [str(cell.value).lower().replace(' ', '_') for cell in header_row if
                         str(cell.value).lower() != 'none']

        # Iterate over the rows within the specified range
        for row_number in range(start_row, end_row + 1):
            for col_idx, column_letter in enumerate(column_letters):
                cell_value = worksheet[column_letter + str(row_number)].value
                if str(cell_value).lower() != 'none':
                    list_of_lists[col_idx].append(cell_value)

        # Close the workbook
        workbook.close()

        return list_of_lists, header_values

    except Exception as e:
        print(f"Error while processing the file '{file_path}': {e}")
        return None, None


def removeNone_listOfLists(listName):
    for i, sublist in enumerate(listName):
        listName[i] = [value for value in sublist if value is not None]
    return listName


def removeEmptyList(listName):
    cleaned_list = [sublist for sublist in listName if sublist]
    return cleaned_list


def remove_stock_symbol(text):
    # Define a regular expression pattern to match stock symbols
    stock_pattern = re.compile(r'\b[A-Z]+\b')
    # Find all stock symbols in the text
    if stock_pattern is not None:
        stock_symbols = stock_pattern.findall(text)
    else:
        stock_symbols = 'not found'

    # return stock value
    return stock_symbols


def remove_and_extract_date(text):
    date_pattern = re.compile(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec) \d{1,2} \d{4}')
    match = date_pattern.search(text)
    if match:
        extracted_date = match.group()
        return extracted_date
    else:
        return 'not found'


def extract_strike_value(text):
    strike_pattern = re.compile(r'\b\d+\.\d+\b')
    match = strike_pattern.search(text)

    if match:
        return float(match.group())
    else:
        return 'not found'


def extract_stock_type(text):
    stock_type_pattern = re.compile(r'\b(Call|Put)\b', re.IGNORECASE)
    match = stock_type_pattern.search(text)

    if match:
        return match.group().capitalize()
    else:
        return 'not found'


def check_buying_situation(text):
    buying_sit_pattern = re.compile(r'\b(Sold|Bought)\b', re.IGNORECASE)
    match = buying_sit_pattern.search(text)
    if match:
        return match.group()
    else:
        return 'not found'


def has_sequence_case_insensitive(my_list, sequence):
    # Convert both the list and the sequence to lowercase
    my_list_lower = [str(item).lower() for item in my_list]
    sequence_lower = [str(item).lower() for item in sequence]

    # Iterate through the modified list and check for the modified sequence
    for i in range(len(my_list_lower) - len(sequence_lower) + 1):
        if my_list_lower[i:i + len(sequence_lower)] == sequence_lower:
            return True
    return False


def convert_date_to_dd_mm_yyyy(input_format, output_format, date_item):
    # Parse the date string into a datetime object
    datetime_object = datetime.strptime(date_item, input_format)
    date_formatted = datetime_object.strftime(output_format)
    # Extract the date part and return it as datetime.date
    return date_formatted


def retrieve_numerical_value(input_text):
    numeric_part = re.search(r'\d+', input_text).group()
    numeric_value = int(numeric_part)
    return numeric_value


def last_five_months_numbers():
    # Get today's date
    today = datetime.now()

    # Calculate the last five months
    last_five_months = [today - timedelta(days=30 * i) for i in range(4)]

    # Extract and print only the month numbers in 'mm' format
    month_numbers = [date.strftime('%m') for date in last_five_months]
    return month_numbers


def find_next_11_fridays():
    # Get today's date
    today = datetime.now().date()

    # Find the next Friday from today
    days_until_next_friday = (4 - today.weekday() + 7) % 7
    next_friday = today + timedelta(days=days_until_next_friday)

    # Calculate the dates for the next 10 Fridays
    next_fridays_primary_list = [next_friday + timedelta(weeks=i) for i in range(11)]
    next_fridays = []
    # Print the result in "mm/dd/yy" format
    for date in next_fridays_primary_list:
        formatted_date = date.strftime("%-m/%d/%y")
        next_fridays.append(formatted_date)

    return next_fridays


def calculate_days_difference(date1, date2):
    result = date1 - date2
    return result.days


def extract_underlying_symbol(data_input):
    u_symbol = []
    header_list_check = ['Date', 'Description', 'Quantity', 'Symbol', 'Price', 'Amount']
    xlsx_file = get_xlsx(data_input)
    if xlsx_file is not None:
        first_sheet_name, file_path, columns, last_row = xlsx_file
        result_lists, headers = get_data_from_range(file_path, first_sheet_name, columns, 2, last_row)
        if has_sequence_case_insensitive(headers, header_list_check):
            # cleaning result list
            new_resultList = removeNone_listOfLists(result_lists)
            final_resultList = removeEmptyList(new_resultList)

            ### - - - - - separating the content of the first list => result: 4 lists - - - - - #
            # remove string content
            final_resultList[0].remove('***END OF FILE***')

            # New Lists: option_expiration_date, strike, underlying_symbol, stock_type
            for n in final_resultList[1]:
                # --- underlying_symbol -- #
                remove_value = remove_stock_symbol(n)
                stock_symbol = ', '.join(remove_value)
                u_symbol.append(stock_symbol)
    return u_symbol


def process_xlsx_trade(data_input):
    # - - - Lists to be used - - - - #
    global combined_filled_lists
    header_list_check = ['Date', 'Description', 'Quantity', 'Symbol', 'Price', 'Amount']
    underlying_symbol = []
    otm_at_time_of_trade = []
    underlying_price_current = []
    otm_current = []
    amount_of_stock_itm_can_be_called = []
    weight = []
    weighted_otm = []
    option_expiration_date = []
    strike = []
    stock_type = []
    quantity = []
    mkt_beta_px_contracts = []
    mkt_price_of_contracts = []
    premium = []
    price = []
    trade_price_percent_notional = []
    annual_yield_at_strike = []
    yield_on_cost_at_trade = []
    yield_at_current_mkt_price_at_trade = []
    trade_date = []
    days_till_exp_date = []
    days_till_exp_date_current = []
    stock_buy_list = []
    cash_if_exercised = []

    # months
    month_1 = []
    month_2 = []
    month_3 = []
    month_4 = []
    month_5 = []

    # fridays
    week_1 = []
    week_2 = []
    week_3 = []
    week_4 = []
    week_5 = []
    week_6 = []
    week_7 = []
    week_8 = []
    week_9 = []
    week_10 = []
    week_11 = []

    underlying_price_at_time_of_trade_temporary = ['1', '2', '3']
    mkt_beta_list_temporary = ['2', '7', '8']

    fridays_list = find_next_11_fridays()
    xlsx_file = get_xlsx(data_input)
    if xlsx_file is not None:
        first_sheet_name, file_path, columns, last_row = xlsx_file
        result_lists, headers = get_data_from_range(file_path, first_sheet_name, columns, 2, last_row)
        if has_sequence_case_insensitive(headers, header_list_check):
            print('File validation ✓')
            print('Reading XLSX file...')
            print('XLSX file processed ✓')
            # cleaning result list
            new_resultList = removeNone_listOfLists(result_lists)
            final_resultList = removeEmptyList(new_resultList)

            ### - - - - - separating the content of the first list => result: 4 lists - - - - - #
            # remove string content
            final_resultList[0].remove('***END OF FILE***')

            # get trade date (1st column)
            for n in final_resultList[0]:
                formatted_date = n.strftime("%-m/%d/%y")
                trade_date.append(formatted_date)

            # New Lists: option_expiration_date, strike, underlying_symbol, stock_type
            for n in final_resultList[1]:
                # --- underlying_symbol -- #
                remove_value = remove_stock_symbol(n)
                stock_symbol = ', '.join(remove_value)
                underlying_symbol.append(stock_symbol)

                # --- option_expiration_date -- #
                expiration_date = remove_and_extract_date(n)
                formatted_date = convert_date_to_dd_mm_yyyy("%b %d %Y", "%-m/%d/%y", expiration_date)
                option_expiration_date.append(formatted_date)

                # --- strike -- #
                strike_value = extract_strike_value(n)
                strike.append(strike_value)

                # --- stock_type -- #
                stock_type_v = extract_stock_type(n)
                stock_type.append(stock_type_v)

                # --- Stock buying situation -- #
                stock_buy = check_buying_situation(n)
                stock_buy_list.append(stock_buy)

            # New list: price
            for y in final_resultList[4]:
                price.append(y)

            # New List: quantity
            for index, y in enumerate(final_resultList[2]):
                if stock_buy_list[index].lower() == 'sold':
                    negative_y = y * -1
                    quantity.append(negative_y)
                else:
                    quantity.append(y)

            # New List: premium
            for y in final_resultList[5]:
                premium.append(y)

            # EXCEL FORMULAS
            length_of_data = len(final_resultList[1])
            for index in range(length_of_data):
                string_index = str(index + 3)

                # values for the month contracted columns
                trade_date_comparison = trade_date[index]
                date_object = datetime.strptime(trade_date_comparison, '%m/%d/%y')

                # Extract the month and format it as 'mm'
                trade_month = date_object.strftime('%m')
                header_months = last_five_months_numbers()

                if trade_month == header_months[0]:
                    month_4.append('=AC' + string_index)
                else:
                    month_4.append('')

                if trade_month == header_months[1]:
                    month_1.append('=AC' + string_index)
                else:
                    month_1.append('')

                if trade_month == header_months[2]:
                    month_2.append('=AC' + string_index)
                else:
                    month_2.append('')

                if trade_month == header_months[3]:
                    month_3.append('=AC' + string_index)
                else:
                    month_3.append('')

                option_date = option_expiration_date[index]

                # get values for the friday cells
                fridays_list = find_next_11_fridays()

                friday_1 = fridays_list[0]
                friday_2 = fridays_list[1]
                friday_3 = fridays_list[2]
                friday_4 = fridays_list[3]
                friday_5 = fridays_list[4]
                friday_6 = fridays_list[5]
                friday_7 = fridays_list[6]
                friday_8 = fridays_list[7]
                friday_9 = fridays_list[8]
                friday_10 = fridays_list[9]
                friday_11 = fridays_list[10]

                if option_date == friday_1:
                    week_1.append('=AH' + string_index)
                else:
                    week_1.append('')

                if option_date == friday_2:
                    week_2.append('=AH' + string_index)
                else:
                    week_2.append('')

                if option_date == friday_3:
                    week_3.append('=AH' + string_index)
                else:
                    week_3.append('')

                if option_date == friday_4:
                    week_4.append('=AH' + string_index)
                else:
                    week_4.append('')

                if option_date == friday_5:
                    week_5.append('=AH' + string_index)
                else:
                    week_5.append('')

                if option_date == friday_6:
                    week_6.append('=AH' + string_index)
                else:
                    week_6.append('')

                if option_date == friday_7:
                    week_7.append('=AH' + string_index)
                else:
                    week_7.append('')

                if option_date == friday_8:
                    week_8.append('=AH' + string_index)
                else:
                    week_8.append('')

                if option_date == friday_9:
                    week_9.append('=AH' + string_index)
                else:
                    week_9.append('')

                if option_date == friday_10:
                    week_10.append('=AH' + string_index)
                else:
                    week_10.append('')

                if option_date == friday_11:
                    week_11.append('=AH' + string_index)
                else:
                    week_11.append('')

                # get days till exp -
                formula_exp_date = '=D' + string_index + '-C' + string_index
                days_till_exp_date.append(formula_exp_date)

                # get days till exp current
                formula_exp_date_current = '=D' + string_index + '-$BE$1'
                days_till_exp_date_current.append(formula_exp_date_current)

                # get otm at time of trade
                formula_get_otm_at_time_of_date = '=(I' + string_index + '-K' + string_index + ')/K' + string_index
                otm_at_time_of_trade.append(formula_get_otm_at_time_of_date)

                # underlying price, current
                formula_underlying_price_current = '=IFERROR(GOOGLEFINANCE("' + underlying_symbol[index] + '"), 59.06)'
                underlying_price_current.append(formula_underlying_price_current)

                # otm, current
                formula_otm_current = '=IF(S' + string_index + '= "Call", (I' + string_index + '-M' + string_index + ')/M' + string_index + ', (M' + string_index + '-I' + string_index + ')/M' + string_index + ')'
                otm_current.append(formula_otm_current)

                # $ amount of stock itm can be called (-) or put (+)
                formula_amount_itm = '=IF(AND(LOWER(S' + string_index + ')="call",N' + string_index + '<0), I' + string_index + '*100*U' + string_index + ', 0)'
                amount_of_stock_itm_can_be_called.append(formula_amount_itm)

                # weight
                formula_weight = '=U' + string_index + '*I' + string_index
                weight.append(formula_weight)

                # weighted otm
                formula_weighted_otm = '=U' + string_index + '*N' + string_index + '*I' + string_index
                weighted_otm.append(formula_weighted_otm)

                # mkt beta * mkt px * contracts
                formula_mkt_beta_px_contracts = '=R' + string_index + '*M' + string_index + '*U' + string_index
                mkt_beta_px_contracts.append(formula_mkt_beta_px_contracts)

                # mkt price * # of contracts
                formula_mkt_price_hashtag_contracts = '=U' + string_index + '*M' + string_index
                mkt_price_of_contracts.append(formula_mkt_price_hashtag_contracts)

                # trade price as percent of notional
                formula_trade_price = '=-AC' + string_index + '/AH' + string_index
                trade_price_percent_notional.append(formula_trade_price)

                # annual yield at strike at time of trade
                formula_annual_yield = '=W' + string_index + '*(365/E' + string_index + ')/I' + string_index
                annual_yield_at_strike.append(formula_annual_yield)

                # yield_on_cost_at_trade ----- THIS IS WAITING FOR CONFIRMATION REG THE FORMULA !! --- INCOMPLETE as of 01/07
                formula_yield_on_cost = ''
                yield_on_cost_at_trade.append(formula_yield_on_cost)

                # yield at current mkt price at trade
                formula_yield_current_mkt = '=X' + string_index + '*(365/F' + string_index + ')/L' + string_index
                yield_at_current_mkt_price_at_trade.append(formula_yield_current_mkt)

                # cash if exercised
                formula_cash_if_exercised = '=IF(LOWER(S' + string_index + ')="put", U' + string_index + '*I' + string_index + '*100,-U' + string_index + '*I' + string_index + '*100)'
                cash_if_exercised.append(formula_cash_if_exercised)

                combined_filled_lists = [
                    trade_date, option_expiration_date, days_till_exp_date, days_till_exp_date_current, strike,
                    underlying_symbol,
                    underlying_price_at_time_of_trade_temporary, otm_at_time_of_trade, underlying_price_current,
                    otm_current,
                    amount_of_stock_itm_can_be_called, weight, weighted_otm, mkt_beta_list_temporary, stock_type,
                    mkt_beta_px_contracts, quantity,
                    mkt_price_of_contracts, price, trade_price_percent_notional, annual_yield_at_strike,
                    yield_at_current_mkt_price_at_trade, premium, month_4, month_1, month_2, month_3,
                    cash_if_exercised, week_1,
                    week_2, week_3, week_4, week_5, week_6, week_7, week_8, week_9, week_10, week_11
                ]

        else:
            print(
                f'INVALID INPUT FILE. The file should have headers with the following titles in this order:{header_list_check}')
            print('Please start the program again!')
            sys.exit()
    return combined_filled_lists


def process_xlsx_orders(data_input):
    trade_date_list = []
    underlying_list = []
    underlying_price_time_of_trade_list = []
    otm_at_time_of_trade = []
    underlying_price_current = []
    otm_current = []
    amount_of_stock_itm_can_be_called = []
    weight = []
    weighted_otm = []
    option_expiration_date_list = []
    order_expiration_date_time_in_force = []
    strike_list = []
    type_list = []
    qty_list = []
    mkt_beta_px_contracts = []
    mkt_price_of_contracts = []
    trade_price_premium = []
    trade_price_percent_notional = []
    annual_yield_at_strike = []
    yield_on_cost_at_trade = []
    yield_at_current_mkt_price_at_trade = []
    trade_date = []
    days_till_exp_date = []
    days_till_exp_date_current = []
    stock_buy_list = []
    premium_list = []
    cash_if_exercised = []

    # months
    month_1 = []
    month_2 = []
    month_3 = []
    month_4 = []
    month_5 = []

    # fridays
    week_1 = []
    week_2 = []
    week_3 = []
    week_4 = []
    week_5 = []
    week_6 = []
    week_7 = []
    week_8 = []
    week_9 = []
    week_10 = []
    week_11 = []

    xlsx_file = get_xlsx_orders(data_input)
    if xlsx_file is not None:
        today_date = datetime.now().strftime('%-m/%d/%y')

        first_sheet_name, file_path, columns, last_row = xlsx_file
        result_lists, headers = get_data_from_range(file_path, first_sheet_name, columns, 2, last_row)
        new_resultList = removeNone_listOfLists(result_lists)
        final_resultList = removeEmptyList(new_resultList)

        # loop through quantity list
        for n in final_resultList[3]:
            qty_list.append(n)

        # loop through symbol list to get multiple values:
        for index, n in enumerate(final_resultList[4]):
            length_of_item = len(n)
            if length_of_item < 6:
                # --- option_expiration_date --
                option_expiration_date_list.append(final_resultList[9][index])

                # --- underlying -- #
                remove_value = remove_stock_symbol(n)
                stock_symbol = ', '.join(remove_value)
                underlying_list.append(stock_symbol)

                # --- strike -- #
                strike_list.append(final_resultList[7][index])

                # --- type -- #
                type_list.append(' ')


            else:
                # --- option_expiration_date -- #
                expiration_date = remove_and_extract_date(n)
                formatted_date = convert_date_to_dd_mm_yyyy("%b %d %Y", "%-m/%d/%y", expiration_date)
                option_expiration_date_list.append(formatted_date)

                # --- strike -- #
                strike_value = extract_strike_value(n)
                strike_list.append(strike_value)

                # --- underlying -- #
                remove_value = remove_stock_symbol(n)
                stock_symbol = ', '.join(remove_value)
                underlying_list.append(stock_symbol)

                # --- type -- #
                stock_type_v = extract_stock_type(n)
                type_list.append(stock_type_v)

        # loop through price list:
        for n in final_resultList[7]:
            trade_price_premium.append(n)
            trade_date_list.append(today_date)

        # loop through time in force list
        for n in final_resultList[9]:
            order_expiration_date_time_in_force.append(n)

        # EXCEL FORMULAS
        length_of_data = len(final_resultList[1])
        for index in range(length_of_data):
            string_index = str(index + 3)

            # check if it's an option or a share
            symbol_list_length = len(final_resultList[4][index])

            # values for the month contracted columns
            trade_date_comparison = trade_date_list[index]
            date_object = datetime.strptime(trade_date_comparison, '%m/%d/%y')

            # Extract the month and format it as 'mm'
            trade_month = date_object.strftime('%m')
            header_months = last_five_months_numbers()

            if trade_month == header_months[0]:
                month_4.append('=AC' + string_index)
            else:
                month_4.append('')

            if trade_month == header_months[1]:
                month_1.append('=AC' + string_index)
            else:
                month_1.append('')

            if trade_month == header_months[2]:
                month_2.append('=AC' + string_index)
            else:
                month_2.append('')

            if trade_month == header_months[3]:
                month_3.append('=AC' + string_index)
            else:
                month_3.append('')

            option_date = option_expiration_date_list[index]

            # get values for the friday cells
            fridays_list = find_next_11_fridays()

            friday_1 = fridays_list[0]
            friday_2 = fridays_list[1]
            friday_3 = fridays_list[2]
            friday_4 = fridays_list[3]
            friday_5 = fridays_list[4]
            friday_6 = fridays_list[5]
            friday_7 = fridays_list[6]
            friday_8 = fridays_list[7]
            friday_9 = fridays_list[8]
            friday_10 = fridays_list[9]
            friday_11 = fridays_list[10]

            if option_date == friday_1:
                week_1.append('=AH' + string_index)
            else:
                week_1.append('')

            if option_date == friday_2:
                week_2.append('=AH' + string_index)
            else:
                week_2.append('')

            if option_date == friday_3:
                week_3.append('=AH' + string_index)
            else:
                week_3.append('')

            if option_date == friday_4:
                week_4.append('=AH' + string_index)
            else:
                week_4.append('')

            if option_date == friday_5:
                week_5.append('=AH' + string_index)
            else:
                week_5.append('')

            if option_date == friday_6:
                week_6.append('=AH' + string_index)
            else:
                week_6.append('')

            if option_date == friday_7:
                week_7.append('=AH' + string_index)
            else:
                week_7.append('')

            if option_date == friday_8:
                week_8.append('=AH' + string_index)
            else:
                week_8.append('')

            if option_date == friday_9:
                week_9.append('=AH' + string_index)
            else:
                week_9.append('')

            if option_date == friday_10:
                week_10.append('=AH' + string_index)
            else:
                week_10.append('')

            if option_date == friday_11:
                week_11.append('=AH' + string_index)
            else:
                week_11.append('')

            # get days till exp -
            formula_exp_date = '=D' + string_index + '-C' + string_index
            days_till_exp_date.append(formula_exp_date)

            # get days till exp current
            formula_exp_date_current = '=D' + string_index + '-$BE$1'
            days_till_exp_date_current.append(formula_exp_date_current)

            # get otm at time of trade
            formula_get_otm_at_time_of_date = '=(I' + string_index + '-K' + string_index + ')/K' + string_index
            otm_at_time_of_trade.append(formula_get_otm_at_time_of_date)

            # underlying price at time of trade
            formula_underlying_price_at_time_of_trade = '=M' + string_index
            underlying_price_time_of_trade_list.append(formula_underlying_price_at_time_of_trade)

            # underlying price, current
            formula_underlying_price_current = '=IFERROR(GOOGLEFINANCE("' + underlying_list[index] + '"), 59.06)'
            underlying_price_current.append(formula_underlying_price_current)

            # otm, current
            if symbol_list_length <= 6:
                otm_current.append(' ')
            else:

                formula_otm_current = '=IF(S' + string_index + '= "Call", (I' + string_index + '-M' + string_index + ')/M' + string_index + ', (M' + string_index + '-I' + string_index + ')/M' + string_index + ')'
                otm_current.append(formula_otm_current)

            # $ amount of stock itm can be called (-) or put (+)
            amount_of_stock_itm_can_be_called.append(' ')

            # mkt price * # of contracts
            formula_mkt_price_hashtag_contracts = '=U' + string_index + '*M' + string_index
            mkt_price_of_contracts.append(formula_mkt_price_hashtag_contracts)

            # trade price as percent of notional
            if symbol_list_length <= 6:
                trade_price_percent_notional.append(' ')
            else:
                formula_trade_price = '=-AC' + string_index + '/AH' + string_index
                trade_price_percent_notional.append(formula_trade_price)

            # annual yield at strike at time of trade
            formula_annual_yield = '=W' + string_index + '*(365/E' + string_index + ')/I' + string_index
            annual_yield_at_strike.append(formula_annual_yield)

            # yield_on_cost_at_trade ----- THIS IS WAITING FOR CONFIRMATION REG THE FORMULA !! --- INCOMPLETE as of 01/07
            formula_yield_on_cost = ''
            yield_on_cost_at_trade.append(formula_yield_on_cost)

            # yield at current mkt price at trade
            formula_yield_current_mkt = '=W' + string_index + '*(365/F' + string_index + ')/L' + string_index
            yield_at_current_mkt_price_at_trade.append(formula_yield_current_mkt)

            # premium
            formula_premium = '=U' + string_index + '*W' + string_index + '*100'
            premium_list.append(formula_premium)

            # cash if exercised
            if symbol_list_length <= 6:
                cash_if_exercised.append(' ')
            else:
                formula_cash_if_exercised = '=IF(LOWER(S' + string_index + ')="put", U' + string_index + '*I' + string_index + '*100,-U' + string_index + '*I' + string_index + '*100)'
                cash_if_exercised.append(formula_cash_if_exercised)

        order_content_filled_list = [
            trade_date_list, option_expiration_date_list, days_till_exp_date, days_till_exp_date_current,
            order_expiration_date_time_in_force, strike_list,
            underlying_list, underlying_price_time_of_trade_list,
            otm_at_time_of_trade, underlying_price_current,
            otm_current,
            amount_of_stock_itm_can_be_called, type_list, qty_list,
            trade_price_premium, trade_price_percent_notional, annual_yield_at_strike,
            yield_at_current_mkt_price_at_trade, premium_list, month_4, month_1, month_2, month_3,
            cash_if_exercised, week_1,
            week_2, week_3, week_4, week_5, week_6, week_7, week_8, week_9, week_10, week_11
        ]

        return order_content_filled_list
